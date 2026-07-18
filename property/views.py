import csv
import json
import uuid
import openpyxl
import base64
import io
import re
from datetime import timedelta
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage
from django.db import transaction
from django.db.models import Q, Prefetch
from django.http import HttpResponse
from django.utils import timezone
from .models import Unit, UnitImages, UnitDocuments, UnitOwner, Property, PropertyBlocks, PropertyImages, PropertyDocuments, PropertyManagmentCompany, PropertyInterest
from user_service.models import PropertyManager, DocumentType, Owner, Tenant
from utilities.decorator import is_request_authenticated
from utilities.helper_functions import prepare_response, generate_property_code, upload_file_to_s3_base64, fetch_s3_presigned_url, get_extension_from_base64, export_to_csv
from utilities import status, constants
from property_management.utils import audit_logs, get_full_property_data, get_property_images, get_lease_status
from lease.models import Lease
from property.models import PropertyType
from rest_framework.decorators import api_view
from .swagger import (
    property_get, property_post, property_put,
    property_blocks_get, property_blocks_post, property_blocks_put,
    property_images_get, property_images_post, property_images_delete,
    property_document_types_get,
    property_documents_get, property_documents_post, property_documents_delete,
    unit_get, unit_post, unit_put,
    unit_images_get, unit_images_post, unit_images_delete,
    unit_document_types_get,
    unit_documents_get, unit_documents_post, unit_documents_delete,
)
from plugins.logger_plugin import get_logger

logger = get_logger(__name__)
def _parse_date(value):
    if not value:
        return None
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            return datetime.strptime(str(value)[:26], fmt)
        except ValueError:
            continue
    return None


def _apply_owner_fields(owner_obj, owner_data):
    """Update owner fields with any non-empty values from owner_data."""
    name = owner_data.get("owner_name") or owner_data.get("name") or ""
    if name:
        first_name, _, last_name = name.partition(" ")
        owner_obj.user.first_name = first_name
        owner_obj.user.last_name = last_name
        owner_obj.user.save(update_fields=["first_name", "last_name"])

    field_map = {
        "email": "email",
        "contact_number": "contact_number",
        "emirates_id": "emirate_id",
        "owner_number": "owner_number",
        "trade_license_number": "trade_license_number",
        "license_number": "license_number",
        "license_issuer": "license_issuer",
        "fax_number": "fax_number",
        "po_box_number": "po_box_number",
    }
    update_fields = []
    for data_key, model_field in field_map.items():
        val = owner_data.get(data_key)
        if val:
            setattr(owner_obj, model_field, val)
            update_fields.append(model_field)

    expiry = _parse_date(owner_data.get("license_expiry_date"))
    if expiry:
        owner_obj.license_expiry_date = expiry
        update_fields.append("license_expiry_date")

    if update_fields:
        Owner.objects.filter(pk=owner_obj.pk).update(
            **{f: getattr(owner_obj, f) for f in update_fields}
        )
    return owner_obj


def _get_or_create_owner(owner_data, created_by):
    owner_id = owner_data.get("owner_id")
    email = owner_data.get("email") or ""

    # 1. Look up by owner_id
    if owner_id:
        owner_obj = Owner.objects.select_related("user").filter(id=owner_id).first()
        if owner_obj:
            return _apply_owner_fields(owner_obj, owner_data)

    # 2. Look up by email
    if email:
        owner_obj = Owner.objects.select_related("user").filter(email=email).first()
        if owner_obj:
            return _apply_owner_fields(owner_obj, owner_data)

    # 3. Create new owner
    name = owner_data.get("owner_name") or owner_data.get("name") or ""
    if not name and not email:
        return None

    first_name, _, last_name = name.partition(" ")
    django_user = User.objects.create(
        username=f"owner_{uuid.uuid4().hex[:8]}",
        email=email or f"owner_{uuid.uuid4().hex[:8]}@units.local",
        first_name=first_name,
        last_name=last_name,
    )
    owner_obj = Owner.objects.create(
        created_by=created_by,
        user=django_user,
        email=email,
        contact_number=owner_data.get("contact_number") or "",
        emirate_id=owner_data.get("emirates_id") or "",
        owner_number=owner_data.get("owner_number") or "",
        trade_license_number=owner_data.get("trade_license_number") or "",
        license_number=owner_data.get("license_number") or "",
        license_expiry_date=_parse_date(owner_data.get("license_expiry_date")),
        license_issuer=owner_data.get("license_issuer") or "",
        fax_number=owner_data.get("fax_number") or "",
        po_box_number=owner_data.get("po_box_number") or "",
    )
    return owner_obj


@property_get
@property_post
@property_put
@api_view(["GET", "POST", "PUT"])
@is_request_authenticated
def property(request):
    user_profile = request.user

    if request.method == "GET":
        property_id = request.GET.get("property_id")
        if property_id:
            prop = Property.objects.filter(id=property_id).first()
            if not prop:
                logger.warning(
                    "PROPERTY_NOT_FOUND | user_id=%d | property_id=%s",
                    request.user.id, property_id,)
                return prepare_response(
                    message=constants.PROPERTY_NOT_FOUND,
                    status=status.HTTP_404_NOT_FOUND
                )
            return prepare_response(content=prop._serialize_property(), status=status.HTTP_200_OK)

        search = request.GET.get("search", "").strip()
        property_type = request.GET.get("property_type", "").strip()
        prop_status = request.GET.get("status", "").strip()
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))
        export = request.GET.get("export", "").strip()

        properties = Property.objects.none()
        # ── PMC login ─────────────────────────────────────────
        pm_profile = PropertyManager.objects.filter(pk=user_profile.pk).select_related('company').first()
        # If company exists, fetch that company’s active properties.
        if pm_profile and pm_profile.company:
            properties = Property.objects.filter(pmc=pm_profile.company,is_active=True).order_by("-id")

        # Otherwise, find company created by logged-in user (PMC admin/owner of company).
        if not properties.exists() and not pm_profile:
            own_company = PropertyManagmentCompany.objects.filter(created_by=user_profile.user,is_active=True).first()

            if own_company:
                properties = Property.objects.filter(pmc=own_company,is_active=True).order_by("-id")
 
        # ── Owner login ───────────────────────────────────────
        if not properties.exists():
            owner_profile = Owner.objects.filter(pk=user_profile.pk).first()
 
            if owner_profile:
                properties = Property.objects.filter(
                    property_blocks__block_towers__unit_owners__owner=owner_profile,
                    is_active=True
                ).distinct().order_by("-id")
 
        # ── Tenant login ──────────────────────────────────────
        if not properties.exists():
            tenant_profile = Tenant.objects.filter(pk=user_profile.pk).first()
            if tenant_profile:
                properties = Property.objects.filter(
                    property_blocks__block_towers__leases__tenant=tenant_profile,
                    property_blocks__block_towers__leases__is_active=True,
                    is_active=True
                ).distinct().order_by("-id")

        if search:
            properties = properties.filter(property_name__icontains=search)
        if property_type:
            properties = properties.filter(property_type__code__in=property_type.split(",")).distinct()
        if prop_status:
            properties = properties.filter(status=prop_status)

        if export == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="properties.csv"'
            writer = csv.writer(response)
            writer.writerow(["Property ID", "Property Name", "Type", "Status", "Address", "Blocks", "Units", "Plot No", "Dewa No"])
            for p in properties:
                s = p._serialize_property()
                property_types = ", ".join(item["value"] for item in s.get("property_type", []))
                writer.writerow([s.get("code"), s.get("property_name"), property_types, s.get("status"), s.get("address_line_1"), s.get("no_of_blocks"), s.get("no_of_units"), s.get("plot_no"), s.get("dewa_no")])
            return response

        total = properties.count()
        start = (page - 1) * page_size
        properties = properties[start:start + page_size]
        logger.info(
            "PROPERTY_LIST_FETCHED | user_id=%d | total_records=%d | page=%d",
            request.user.id, total, page )
        return prepare_response(
            content=[p._serialize_property() for p in properties],
            pagination={"total_records": total, "page": page, "page_size": page_size},
            status=status.HTTP_200_OK
        )
    
    elif request.method == "POST":
        data = json.loads(request.body)

        property_name = data.get("property_name")
        if not property_name:
            logger.warning(
                "PROPERTY_CREATE_FAILED | user_id=%d | reason=PROPERTY_NAME_REQUIRED",
                request.user.id,)
            return prepare_response(
                message=constants.PROPERTY_NAME_REQUIRED,
                status=status.HTTP_400_BAD_REQUEST
            )

        pm_profile = PropertyManager.objects.filter(pk=user_profile.pk).select_related('company').first()
        pmc = pm_profile.company if pm_profile else None
        if not pmc:
            logger.warning(
                "PROPERTY_CREATE_DENIED | user_id=%d | reason=NOT_VERIFIED_PROPERTY_MANAGER",
                request.user.id,)
            return prepare_response(
                message=constants.NOT_VERIFIED_PROPERTY_MANAGER,
                status=status.HTTP_403_FORBIDDEN
            )

        prop = Property.objects.create(
            created_by=user_profile.user,
            property_name=property_name,
            #property_type=data.get("property_type") or constants.APARTMENT,
            no_of_blocks=data.get("no_of_blocks") or 0,
            no_of_units=data.get("no_of_units") or 0,
            land_area=data.get("land_area"),
            land_area_unit=data.get("land_area_unit") or constants.SQ_FT,
            land_dm_no=data.get("land_dm_no"),
            plot_no=data.get("plot_no"),
            dewa_no=data.get("dewa_no"),
            address_line_1=data.get("address_line_1") or '',
            address_line_2=data.get("address_line_2") or '',
            landmark=data.get("landmark") or '',
            pincode=data.get("pincode") or '',
            latitude=data.get("latitude"),
            longitude=data.get("longitude"),
            map_address=data.get("map_address"),
            approx_rent=data.get("approx_rent"),
            pmc=pmc,
        )
        property_types = data.get("property_type")
        if not property_types:
            property_types = [
                    {
                        "key": constants.APARTMENT,
                        "value": "Apartment"
                    }
            ]
        codes = [
            item.get("key")
            for item in property_types
            if item.get("key")
        ]
        property_type_objs = PropertyType.objects.filter(code__in=codes)
        prop.property_type.set(property_type_objs)
        logger.info(
            "PROPERTY_CREATED | user_id=%d | property_id=%d | property_name=%s | status=SUCCESS",
            request.user.id, prop.id, prop.property_name,)

        audit_logs(request, f"Property '{prop.property_name}' created", constants.CREATED)
        return prepare_response(
            message=constants.PROPERTY_ADDED,
            content={"id": prop.id},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "PUT":
        data = json.loads(request.body)
        property_id = data.get("property_id")
        if not property_id:
            logger.warning(
                "PROPERTY_UPDATE_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",
                request.user.id,)
            return prepare_response(
                message=constants.PROPERTY_ID_REQUIRED,
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()
        if not prop:
            logger.warning(
                "PROPERTY_UPDATE_FAILED | user_id=%d | property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, property_id, )
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        for field in ["property_name", "no_of_blocks", "no_of_units",
                      "land_area", "land_area_unit", "land_dm_no", "plot_no",
                      "dewa_no", "address_line_1", "address_line_2", "landmark", "pincode",
                      "latitude", "longitude", "map_address", "approx_rent"]:
            if field in data and data[field] is not None:
                setattr(prop, field, data[field])

        pm_profile = PropertyManager.objects.filter(pk=user_profile.pk).select_related('company').first()
        if pm_profile and pm_profile.company:
            prop.pmc = pm_profile.company

        prop.save()
        property_types = data.get("property_type")
        if property_types is not None:
            codes = [
                item.get("key")
                for item in property_types
                if item.get("key")
            ]
            types = PropertyType.objects.filter(code__in=codes)
            prop.property_type.set(types)
        logger.info(
            "PROPERTY_UPDATED | user_id=%d | property_id=%d | property_name=%s | status=SUCCESS",
            request.user.id, prop.id, prop.property_name,)

        audit_logs(request, f"Parent property '{prop.property_name}' updated", constants.UPDATED)
        return prepare_response(
            message=constants.PROPERTY_UPDATE_SUCCESS,
            content={"id": prop.id},
            status=status.HTTP_200_OK
        )
    else:
        logger.warning(
            "PROPERTY_API_FAILED | user_id=%d | method=%s | reason=METHOD_NOT_ALLOWED",
            request.user.id, request.method )
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

@is_request_authenticated
def property_type_list(request):
    if request.method == "GET":
        data = [
            {
                "key": pt.code,
                "value": pt.name
            }
            for pt in PropertyType.objects.filter(is_active=True).order_by("name")
        ]

        return prepare_response(
            content=data,
            status=status.HTTP_200_OK
        )

@property_blocks_get
@property_blocks_post
@property_blocks_put
@api_view(["GET", "POST", "PUT"])
@is_request_authenticated
def property_blocks(request):
    user_profile = request.user

    if request.method == "GET":
        property_id = request.GET.get("property_id")
        if not property_id:
            logger.warning(
                "BLOCK_FETCH_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",
                request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        blocks = PropertyBlocks.objects.filter(property_id=property_id)
        content = [
            {
                "id": b.id,
                "block_name": b.block_name,
                "makani_no": b.makani_no,
                "no_of_floors": b.no_of_floors,
                "no_of_parking": b.no_of_parking,
                "no_of_units": b.no_of_units,
            }
            for b in blocks
        ]
        logger.info(
            "BLOCKS_FETCHED | user_id=%d | property_id=%s | block_count=%d",
            request.user.id, property_id, len(content) )
        return prepare_response(content=content, status=status.HTTP_200_OK)

    elif request.method == "POST":
        data = json.loads(request.body)
        property_id = data.get("property_id")
        blocks_data = data.get("blocks") or []

        if not property_id:
            logger.warning(
                "BLOCK_CREATE_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",
                request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()
        if not prop:
            logger.warning(
                "BLOCK_CREATE_FAILED | user_id=%d | property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, property_id,)
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        created = []
        for block in blocks_data:
            b = PropertyBlocks.objects.create(
                created_by=user_profile.user,
                property=prop,
                block_name=block.get("block_name", ""),
                makani_no=block.get("makani_no") or "",
                no_of_floors=block.get("no_of_floors") or 0,
                no_of_parking=block.get("no_of_parking") or 0,
                no_of_units=block.get("no_of_units") or 0,
            )
            created.append(b.id)
        logger.info(
            "BLOCKS_CREATED | user_id=%d | property_id=%d | status=SUCCESS",
            request.user.id, prop.id,)

        return prepare_response(
            message="Blocks added successfully",
            content={"ids": created},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "PUT":
        data = json.loads(request.body)
        property_id = data.get("property_id")
        blocks_data = data.get("blocks") or []

        if not property_id:
            logger.warning(
                "BLOCK_UPDATE_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",
                request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()
        if not prop:
            logger.warning(
                "BLOCK_UPDATE_FAILED | user_id=%d | property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, property_id,)
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        # Replace all blocks for this property
        PropertyBlocks.objects.filter(property=prop).delete()
        for block in blocks_data:
            PropertyBlocks.objects.create(
                created_by=user_profile.user,
                property=prop,
                block_name=block.get("block_name", ""),
                makani_no=block.get("makani_no") or "",
                no_of_floors=block.get("no_of_floors") or 0,
                no_of_parking=block.get("no_of_parking") or 0,
                no_of_units=block.get("no_of_units") or 0,
            )
        logger.info(
            "BLOCKS_UPDATED_SUCCESSFULLY | user_id=%d | property_id=%d | status=SUCCESS",
            request.user.id, prop.id, )

        return prepare_response(
            message="Blocks updated successfully",
            status=status.HTTP_200_OK
        )

    else:
        logger.warning(
            "BLOCK_API_FAILED | user_id=%d | method=%s | reason=METHOD_NOT_ALLOWED", 
            request.user.id, request.method )
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@property_images_get
@property_images_post
@property_images_delete
@api_view(["GET", "POST", "DELETE"])
@is_request_authenticated
def property_images(request):
    user_profile = request.user

    if request.method == "GET":
        property_id = request.GET.get("property_id")
        if not property_id:
            logger.warning(
                "IMAGE_FETCH_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        images = PropertyImages.objects.filter(property_id=property_id)
        content = [
            {
                "id": img.id,
                "file_name": img.file_name,
                "image_type": img.image_type,
                "url": fetch_s3_presigned_url(img.image_path, img.file_name),
            }
            for img in images
        ]
        return prepare_response(content=content, status=status.HTTP_200_OK)

    elif request.method == "POST":
        data = json.loads(request.body)
        property_id = data.get("property_id")
        images_data = data.get("images") or []

        if not property_id:
            logger.warning(
                "IMAGE_UPLOAD_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()
        if not prop:
            logger.warning(
                "IMAGE_UPLOAD_FAILED | user_id=%d | property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, property_id, )
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        company_id = prop.pmc.id if prop.pmc else "unknown"
        created = []
        for img_data in images_data:
            base64_data = img_data.get("data")
            if not base64_data:
                continue
            file_name = img_data.get("file_name") or f"{uuid.uuid4()}.jpg"
            image_type = img_data.get("type") or "INTERIOR"
            ext = get_extension_from_base64(base64_data) or ".jpg"
            unique_filename = f"{uuid.uuid4()}{ext}"
            s3_key = f"company/{company_id}/property/{property_id}/{image_type.lower()}/{unique_filename}"
            url = upload_file_to_s3_base64(base64_data, s3_key)
            img = PropertyImages.objects.create(
                created_by=user_profile.user,
                property=prop,
                image_path=url,
                image_type=image_type,
                file_name=file_name,
            )
            created.append(img.id)
        logger.info(
            "IMAGES_UPLOADED | user_id=%d | property_id=%d | image_count=%d | status=SUCCESS",
            request.user.id, prop.id, len(created), )
        return prepare_response(
            message="Images uploaded successfully",
            content={"ids": created},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "DELETE":
        image_id = request.GET.get("image_id")
        if not image_id:
            logger.warning(
                "IMAGE_DELETE_FAILED | user_id=%d | reason=IMAGE_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="image_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        PropertyImages.objects.filter(id=image_id).delete()
        logger.info(
            "IMAGE_DELETED | user_id=%d | image_id=%d | status=SUCCESS",request.user.id,int(image_id),)
        return prepare_response(
            message="Image deleted successfully",
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@property_document_types_get
@api_view(["GET"])
@is_request_authenticated
def property_document_types(request):
    """GET /property/document-types — returns DocumentType records for the PROPERTY section."""
    if request.method != "GET":
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    doc_types = DocumentType.objects.filter(section=constants.PROPERTY).order_by("id")
    content = [{"id": dt.id, "name": dt.name} for dt in doc_types]
    return prepare_response(content=content, status=status.HTTP_200_OK)


@property_documents_get
@property_documents_post
@property_documents_delete
@api_view(["GET", "POST", "DELETE"])
@is_request_authenticated
def property_documents(request):
    user_profile = request.user

    if request.method == "GET":
        property_id = request.GET.get("property_id")
        if not property_id:
            logger.warning(
                "DOCUMENT_FETCH_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        docs = PropertyDocuments.objects.filter(property_id=property_id).select_related("document_type")
        content = [
            {
                "id": doc.id,
                "file_name": doc.file_name,
                "document_type_id": doc.document_type.id,
                "document_type_name": doc.document_type.name,
                "url": fetch_s3_presigned_url(doc.file_path, doc.file_name),
            }
            for doc in docs
        ]
        logger.info(
            "DOCUMENTS_FETCHED | user_id=%d | property_id=%s | document_count=%d",
            request.user.id, property_id, len(content) )
        return prepare_response(content=content, status=status.HTTP_200_OK)

    elif request.method == "POST":
        data = json.loads(request.body)
        property_id = data.get("property_id")
        documents_data = data.get("documents") or []

        if not property_id:
            logger.warning(
                "DOCUMENT_UPLOAD_FAILED | user_id=%d | reason=PROPERTY_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()
        if not prop:
            logger.warning(
                "DOCUMENT_UPLOAD_FAILED | user_id=%d | property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, property_id, )
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        company_id = prop.pmc.id if prop.pmc else "unknown"
        created = []
        for doc_data in documents_data:
            base64_data = doc_data.get("data")
            if not base64_data:
                continue
            file_name = doc_data.get("file_name") or f"{uuid.uuid4()}.pdf"
            document_type_id = doc_data.get("document_type_id")
            doc_type = DocumentType.objects.filter(id=document_type_id).first() if document_type_id else None
            if not doc_type:
                continue
            ext = get_extension_from_base64(base64_data) or ".pdf"
            unique_filename = f"{uuid.uuid4()}{ext}"
            s3_key = f"company/{company_id}/property/{property_id}/documents/{doc_type.id}/{unique_filename}"
            url = upload_file_to_s3_base64(base64_data, s3_key)
            doc = PropertyDocuments.objects.create(
                created_by=user_profile.user,
                property=prop,
                document_type=doc_type,
                file_name=file_name,
                file_path=url,
            )
            created.append(doc.id)
        logger.info(
            "DOCUMENTS_UPLOADED | user_id=%d | property_id=%d | document_count=%d | status=SUCCESS",
            request.user.id, prop.id, len(created),)
        return prepare_response(
            message="Documents uploaded successfully",
            content={"ids": created},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "DELETE":
        document_id = request.GET.get("document_id")
        if not document_id:
            logger.warning(
                "DOCUMENT_DELETE_FAILED | user_id=%d | reason=DOCUMENT_ID_REQUIRED", request.user.id,)
            return prepare_response(
                message="document_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        PropertyDocuments.objects.filter(id=document_id).delete()
        logger.info(
            "DOCUMENT_DELETED | user_id=%d | document_id=%d | status=SUCCESS",
            request.user.id, int(document_id), )
        return prepare_response(
            message="Document deleted successfully",
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@unit_get
@unit_post
@unit_put
@api_view(["GET", "POST", "PUT"])
@is_request_authenticated
def unit(request):
    user_profile = request.user

    if request.method == "GET":
        unit_id = request.GET.get("unit_id")
        if unit_id:
            u = Unit.objects.filter(id=unit_id).select_related(
                # "property_block_tower__property"
                "property_block_tower__property", "parent_property"
            ).prefetch_related("unit_owners").first()
            if not u:
                logger.warning(
                    "UNIT_NOT_FOUND | user_id=%d | unit_id=%s", request.user.id, unit_id, )
                return prepare_response(
                    message=constants.UNIT_NOT_FOUND,
                    status=status.HTTP_404_NOT_FOUND
                )
            logger.info(
                "UNIT_FETCHED | user_id=%d | unit_id=%d",
                request.user.id, u.id )
            return prepare_response(content=u._serialize_unit(), status=status.HTTP_200_OK)

        search = request.GET.get("search", "").strip()
        property_id = request.GET.get("property_id", "").strip()
        block_id = request.GET.get("block_id", "").strip()
        no_of_bedrooms = request.GET.get("no_of_bedrooms", "").strip()
        floor_no = request.GET.get("floor_no", "").strip()
        land_area_unit = request.GET.get("land_area_unit", "").strip()
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 10))
        export = request.GET.get("export", "").strip()

        # ── Role detection: PMC → Owner → Tenant ────────────────
        pm_profile = PropertyManager.objects.filter(pk=user_profile.pk).select_related("company").first()
        if pm_profile and pm_profile.company:
            company = pm_profile.company

            if not company:
                company = PropertyManagmentCompany.objects.filter(
                    created_by=user_profile.user, is_active=True
                ).first()

            if not company:
                logger.warning(
                "UNIT_FETCH_FAILED | user_id=%d | reason=COMPANY_NOT_FOUND",request.user.id,)
                return prepare_response(
                    message=constants.COMPANY_NOT_FOUND,
                    status=status.HTTP_404_NOT_FOUND
                )

            units = Unit.objects.filter(
                Q(property_block_tower__property__pmc=company) |
                Q(parent_property__pmc=company)
            ).select_related(
                "property_block_tower__property",
                "parent_property"
            ).prefetch_related("unit_owners").order_by("-id")

        else:
            owner_profile = Owner.objects.filter(pk=user_profile.pk).first()

            if owner_profile:
                units = Unit.objects.filter(
                    unit_owners__owner=owner_profile,
                    is_active=True
                ).select_related(
                    "property_block_tower__property"
                ).prefetch_related("unit_owners").distinct().order_by("-id")
 
            else:
                tenant_profile = Tenant.objects.filter(pk=user_profile.pk).first() 
                if tenant_profile:
                    units = Unit.objects.filter(
                        leases__tenant=tenant_profile,
                        leases__is_active=True,
                        is_active=True
                    ).select_related(
                        "property_block_tower__property"
                    ).prefetch_related("unit_owners").distinct().order_by("-id") 
                else:
                    return prepare_response(
                        message=constants.UNAUTHORIZED_ROLE,
                        status=status.HTTP_403_FORBIDDEN
                    )

        if search:
            units = units.filter(unit_name__icontains=search)
        if property_id:
            # units = units.filter(property_block_tower__property_id=property_id)
            units = units.filter(Q(property_block_tower__property_id=property_id) | Q(parent_property_id=property_id))
        if block_id:
            units = units.filter(property_block_tower_id=block_id)
        if no_of_bedrooms:
            units = units.filter(no_of_bedrooms=no_of_bedrooms)
        if floor_no:
            units = units.filter(floor_no=floor_no)
        if land_area_unit:
            # units = units.filter(
            #     property_block_tower__property__land_area_unit=land_area_unit.upper())
            units = units.filter(
                Q(property_block_tower__property__land_area_unit=land_area_unit.upper())
                | Q(parent_property__land_area_unit=land_area_unit.upper())
            )

        if export == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="units.csv"'
            writer = csv.writer(response)
            writer.writerow(["Unit ID", "Unit Name", "Property", "Block/Tower", "Bedrooms", "Floor No", "Land No", "Unit Size", "Area", "Makani No", "Dewa No", "Rent (AED)"])
            for u in units:
                s = u._serialize_unit()
                writer.writerow([s.get("code"), s.get("unit_name"), s.get("property_name"), s.get("block_name"), s.get("no_of_bedrooms"), s.get("floor_no"), s.get("land_no"), s.get("unit_size"), f"{s.get('area') or ''} {s.get('land_area_unit') or ''}".strip(), s.get("makani_no"), s.get("dewa_no"), s.get("rent")])
            return response

        total = units.count()
        start = (page - 1) * page_size
        units = units[start:start + page_size]
        logger.info(
            "UNIT_LIST_FETCHED | user_id=%d | total_records=%d | page=%d",
            request.user.id, total, page )
        return prepare_response(
            content=[u._serialize_unit() for u in units],
            pagination={"total_records": total, "page": page, "page_size": page_size},
            status=status.HTTP_200_OK
        )

    elif request.method == "POST":
        data = json.loads(request.body)
        block_id = data.get("block_id")
        parent_property_id = data.get("parent_property_id") or data.get("property_id")
        unit_name = data.get("unit_name")

        # if not block_id:
        if not block_id and not parent_property_id:
            logger.warning(
                "UNIT_CREATE_FAILED | user_id=%d | reason=PROPERTY_OR_BLOCK_REQUIRED", request.user.id,)
            return prepare_response(
                # message="block_id is required",
                message="block_id or parent_property_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        if not unit_name:
            logger.warning(
                "UNIT_CREATE_FAILED | user_id=%d | reason=UNIT_NAME_REQUIRED", request.user.id,)
            return prepare_response(
                message="unit_name is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        # block = PropertyBlocks.objects.filter(id=block_id).first()
        block = PropertyBlocks.objects.filter(id=block_id).first() if block_id else None
        # if not block:
        if block_id and not block:
            logger.warning(
                "UNIT_CREATE_FAILED | user_id=%d | block_id=%s | reason=BLOCK_NOT_FOUND",
                request.user.id, block_id, )
            return prepare_response(
                message="Block not found",
                status=status.HTTP_404_NOT_FOUND
            )
        # parent_property = block.property
        parent_property = block.property if block else Property.objects.filter(id=parent_property_id).first()
        if not parent_property:
            logger.warning(
                "UNIT_CREATE_FAILED | user_id=%d | parent_property_id=%s | reason=PROPERTY_NOT_FOUND",
                request.user.id, parent_property_id, )
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        u = Unit.objects.create(
            created_by=user_profile.user,
            # parent_property=block.property,
            parent_property=parent_property,
            property_block_tower=block,
            unit_name=unit_name,
            unit_size=data.get("unit_size"),
            area=data.get("area"),
            dm_no=data.get("dm_no"),
            no_of_bedrooms=data.get("no_of_bedrooms"),
            floor_no=data.get("floor_no"),
            parking_no=data.get("parking_no"),
            no_of_balcony=data.get("no_of_balcony"),
            land_no=data.get("land_no"),
            unit_usage=data.get("unit_usage"),
            unit_type=data.get("unit_type"),
            sub_type=data.get("sub_type"),
            makani_no=data.get("makani_no"),
            dewa_no=data.get("dewa_no"),
            rent=data.get("rent"),
            security_deposit=data.get("security_deposit"),
            booking_amount=data.get("booking_amount"),
            maintenance_charges=data.get("maintenance_charges"),
            cycle=data.get("cycle"),
            notice_period=data.get("notice_period"),
            commission_percent=data.get("commission_percent"),
        )
        logger.info(
            "UNIT_CREATED | user_id=%d | unit_id=%d | unit_name=%s | status=SUCCESS",
            request.user.id, u.id, u.unit_name, )

        for owner in data.get("unit_owners", []):
            owner_obj = _get_or_create_owner(owner, user_profile.user)
            if owner_obj:
                UnitOwner.objects.create(
                    created_by=user_profile.user,
                    unit=u,
                    owner=owner_obj,
                )

        audit_logs(request, f"Unit '{u.unit_name}' created", constants.CREATED)
        return prepare_response(
            message="Unit added successfully",
            content={"id": u.id},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "PUT":
        data = json.loads(request.body)
        unit_id = data.get("unit_id")
        if not unit_id:
            logger.warning(
                "UNIT_UPDATE_FAILED | user_id=%d | reason=UNIT_ID_REQUIRED", request.user.id, )
            return prepare_response(
                message="unit_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        u = Unit.objects.filter(id=unit_id).first()
        if not u:
            logger.warning(
                "UNIT_UPDATE_FAILED | user_id=%d | unit_id=%s | reason=UNIT_NOT_FOUND",
                request.user.id, unit_id, )
            return prepare_response(
                message=constants.UNIT_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        for field in ["unit_name", "unit_size", "area", "dm_no",
                      "no_of_bedrooms", "floor_no", "parking_no", "no_of_balcony",
                      "land_no", "unit_usage", "unit_type", "sub_type",
                      "makani_no", "dewa_no", "rent", "security_deposit",
                      "booking_amount", "maintenance_charges", "cycle",
                      "notice_period", "commission_percent"]:
            if field in data and data[field] is not None:
                setattr(u, field, data[field])

        if data.get("block_id"):
            block = PropertyBlocks.objects.filter(id=data["block_id"]).first()
            if block:
                u.property_block_tower = block
                u.parent_property= block.property
        # elif data.get("parent_property_id") or data.get("property_id"):
        elif "parent_property_id" in data or "property_id" in data:
            parent_property_id = data.get("parent_property_id") or data.get("property_id")
            parent_property = Property.objects.filter(id=parent_property_id).first()
            if parent_property:
                u.property_block_tower = None
                u.parent_property = parent_property

        u.save()
        logger.info(
            "UNIT_UPDATED | user_id=%d | unit_id=%d | unit_name=%s | status=SUCCESS",
            request.user.id, u.id, u.unit_name, )

        if "unit_owners" in data:
            u.unit_owners.all().delete()
            for owner in data["unit_owners"]:
                owner_obj = _get_or_create_owner(owner, user_profile.user)
                if owner_obj:
                    UnitOwner.objects.create(
                        created_by=user_profile.user,
                        unit=u,
                        owner=owner_obj,
                    )

        audit_logs(request, f"Unit '{u.unit_name}' updated", constants.UPDATED)
        return prepare_response(
            message="Unit updated successfully",
            content={"id": u.id},
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@unit_images_get
@unit_images_post
@unit_images_delete
@api_view(["GET", "POST", "DELETE"])
@is_request_authenticated
def unit_images(request):
    user_profile = request.user

    if request.method == "GET":
        unit_id = request.GET.get("unit_id")
        if not unit_id:
            logger.warning(
                "UNIT_IMAGE_FETCH_FAILED | user_id=%d | reason=UNIT_ID_REQUIRED", request.user.id, )
            return prepare_response(
                message="unit_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        images = UnitImages.objects.filter(unit_id=unit_id)
        content = [
            {
                "id": img.id,
                "file_name": img.file_name,
                "image_type": img.image_type,
                "url": fetch_s3_presigned_url(img.image_path, img.file_name),
            }
            for img in images
        ]
        return prepare_response(content=content, status=status.HTTP_200_OK)

    elif request.method == "POST":
        data = json.loads(request.body)
        unit_id = data.get("unit_id")
        images_data = data.get("images") or []

        if not unit_id:
            logger.warning(
                "UNIT_IMAGE_UPLOAD_FAILED | user_id=%d | reason=UNIT_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="unit_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        u = Unit.objects.filter(id=unit_id).select_related(
            # "property_block_tower__property"
            "property_block_tower__property", "parent_property"
        ).first()
        if not u:
            logger.warning(
                "UNIT_IMAGE_UPLOAD_FAILED | user_id=%d | unit_id=%s | reason=UNIT_NOT_FOUND",
                request.user.id, unit_id, )
            return prepare_response(
                message=constants.UNIT_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        # property_id = u.property_block_tower.property_id
        property_id = u.property_block_tower.property_id if u.property_block_tower_id else u.parent_property_id
        created = []
        for img_data in images_data:
            base64_data = img_data.get("data")
            if not base64_data:
                continue
            file_name = img_data.get("file_name") or f"{uuid.uuid4()}.jpg"
            image_type = img_data.get("type") or "INTERIOR"
            ext = get_extension_from_base64(base64_data) or ".jpg"
            unique_filename = f"{uuid.uuid4()}{ext}"
            s3_key = f"property/{property_id}/unit/{unit_id}/{image_type.lower()}/{unique_filename}"
            url = upload_file_to_s3_base64(base64_data, s3_key)
            img = UnitImages.objects.create(
                created_by=user_profile.user,
                unit=u,
                image_path=url,
                image_type=image_type,
                file_name=file_name,
            )
            created.append(img.id)
        logger.info(
            "UNIT_IMAGES_UPLOADED | user_id=%d | unit_id=%d | image_count=%d | status=SUCCESS",
            request.user.id, u.id, len(created), )
        return prepare_response(
            message="Images uploaded successfully",
            content={"ids": created},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "DELETE":
        image_id = request.GET.get("image_id")
        if not image_id:
            logger.warning(
                "UNIT_IMAGE_DELETE_FAILED | user_id=%d | reason=IMAGE_ID_REQUIRED",request.user.id, )
            return prepare_response(
                message="image_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        UnitImages.objects.filter(id=image_id).delete()
        logger.info(
            "UNIT_IMAGE_DELETED | user_id=%d | image_id=%d | status=SUCCESS",
            request.user.id, int(image_id),)
        return prepare_response(
            message="Image deleted successfully",
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@unit_document_types_get
@api_view(["GET"])
@is_request_authenticated
def unit_document_types(request):
    if request.method != "GET":
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    doc_types = DocumentType.objects.filter(section=constants.UNIT).order_by("id")
    content = [{"id": dt.id, "name": dt.name} for dt in doc_types]
    return prepare_response(content=content, status=status.HTTP_200_OK)


@unit_documents_get
@unit_documents_post
@unit_documents_delete
@api_view(["GET", "POST", "DELETE"])
@is_request_authenticated
def unit_documents(request):
    user_profile = request.user

    if request.method == "GET":
        unit_id = request.GET.get("unit_id")
        if not unit_id:
            logger.warning(
                "UNIT_DOCUMENT_FETCH_FAILED | user_id=%d | reason=UNIT_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="unit_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        docs = UnitDocuments.objects.filter(unit_id=unit_id).select_related("document_type")
        content = [
            {
                "id": doc.id,
                "file_name": doc.file_name,
                "document_type_id": doc.document_type.id,
                "document_type_name": doc.document_type.name,
                "url": fetch_s3_presigned_url(doc.file_path, doc.file_name),
            }
            for doc in docs
        ]
        return prepare_response(content=content, status=status.HTTP_200_OK)

    elif request.method == "POST":
        data = json.loads(request.body)
        unit_id = data.get("unit_id")
        documents_data = data.get("documents") or []

        if not unit_id:
            logger.warning(
                "UNIT_DOCUMENT_UPLOAD_FAILED | user_id=%d | reason=UNIT_ID_REQUIRED",request.user.id,)
            return prepare_response(
                message="unit_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )

        u = Unit.objects.filter(id=unit_id).select_related(
            # "property_block_tower__property"
            "property_block_tower__property", "parent_property"
        ).first()
        if not u:
            logger.warning(
                "UNIT_DOCUMENT_UPLOAD_FAILED | user_id=%d | unit_id=%s | reason=UNIT_NOT_FOUND",
                request.user.id, unit_id, )
            return prepare_response(
                message=constants.UNIT_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        # property_id = u.property_block_tower.property_id
        property_id = u.property_block_tower.property_id if u.property_block_tower_id else u.parent_property_id
        created = []
        for doc_data in documents_data:
            base64_data = doc_data.get("data")
            if not base64_data:
                continue
            file_name = doc_data.get("file_name") or f"{uuid.uuid4()}.pdf"
            document_type_id = doc_data.get("document_type_id")
            doc_type = DocumentType.objects.filter(id=document_type_id).first() if document_type_id else None
            if not doc_type:
                continue
            ext = get_extension_from_base64(base64_data) or ".pdf"
            unique_filename = f"{uuid.uuid4()}{ext}"
            s3_key = f"property/{property_id}/unit/{unit_id}/documents/{doc_type.id}/{unique_filename}"
            url = upload_file_to_s3_base64(base64_data, s3_key)
            doc = UnitDocuments.objects.create(
                created_by=user_profile.user,
                unit=u,
                document_type=doc_type,
                file_name=file_name,
                file_path=url,
            )
            created.append(doc.id)
        logger.info(
            "UNIT_DOCUMENTS_UPLOADED | user_id=%d | unit_id=%d | document_count=%d | status=SUCCESS",
            request.user.id, u.id, len(created), )
        return prepare_response(
            message="Documents uploaded successfully",
            content={"ids": created},
            status=status.HTTP_201_CREATED
        )

    elif request.method == "DELETE":
        document_id = request.GET.get("document_id")
        if not document_id:
            logger.warning(
                "UNIT_DOCUMENT_DELETE_FAILED | user_id=%d | reason=DOCUMENT_ID_REQUIRED",request.user.id, )
            return prepare_response(
                message="document_id is required",
                status=status.HTTP_400_BAD_REQUEST
            )
        UnitDocuments.objects.filter(id=document_id).delete()
        logger.info(
            "UNIT_DOCUMENT_DELETED | user_id=%d | document_id=%d | status=SUCCESS",
            request.user.id, int(document_id),)
        return prepare_response(
            message="Document deleted successfully",
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


# ---------------------------------------------------------------------------
# Functions moved from property_management/views.py
# ---------------------------------------------------------------------------

@is_request_authenticated
def property_table_view(request):
    user = request.user
    search = request.GET.get("search", "").strip()
    page = int(request.GET.get("page", 1))
    limit = int(request.GET.get("limit", 10))
    property_id = request.GET.get("property_id")
    my_property = request.GET.get("MY_PROPERTY", "").lower() == "true"
    tenancy_status = request.GET.get("tenancy_status")
    agreement_expiration = request.GET.get("agreement_expiration")

    if my_property and user.user_role == constants.TENANT:
        lease_property_id = (
        Lease.objects
        .filter(tenant=user)
        .values_list("lease_property_id", flat=True)
        .first()
    )
        if not lease_property_id:
            logger.warning(
                "PROPERTY_TABLE_FETCH_FAILED | user_id=%d | reason=NO_PROPERTY_ASSIGNED",request.user.id, )
            return prepare_response(
            message=constants.NO_PROPERTY_ASSIGNED_TO_TENANAT,
            status=status.HTTP_404_NOT_FOUND
        )
        full_data, error = get_full_property_data(lease_property_id)
        if error:
            logger.warning(
                "PROPERTY_DETAILS_FETCH_FAILED | user_id=%d | property_id=%s | reason=%s",
                request.user.id, lease_property_id, error,)
            return prepare_response(message=error, status=status.HTTP_404_NOT_FOUND)
        return prepare_response(
        content=full_data,
        message=constants.PROPERTIES_FETCHED,
        status=status.HTTP_200_OK
    )

    if property_id:
        full_data, error = get_full_property_data(property_id)
        if error:
            logger.warning(
                "PROPERTY_DETAILS_FETCH_FAILED | user_id=%d | property_id=%s | reason=%s",
                request.user.id, property_id, error,)
            return prepare_response(message=error, status=status.HTTP_404_NOT_FOUND)
        return prepare_response(content=full_data, message=constants.PROPERTIES_FETCHED, status=status.HTTP_200_OK)
    if user.user_role == constants.OWNER:
        properties_qs = Unit.objects.filter(owner=user)

    elif user.user_role == constants.COMPANY_USER:
        companies_qs = PropertyManagmentCompany.objects.filter(company_user=user)

        if not companies_qs.exists():
            logger.warning(
                "PROPERTY_TABLE_FETCH_FAILED | user_id=%d | reason=COMPANY_NOT_FOUND",request.user.id, )
            return prepare_response(
            message=constants.COMPANY_NOT_FOUND,
            status=status.HTTP_400_BAD_REQUEST
        )
        properties_qs = Unit.objects.filter(
        company__in=companies_qs
    )

    elif user.user_role == constants.TENANT:
        # properties_qs = PropertyUnit.objects.filter(lease_details__tenant=user)
        properties_qs = Unit.objects.filter(is_occupied=False)
    else:
        logger.warning(
            "PROPERTY_TABLE_FETCH_DENIED | user_id=%d | role=%s",
            request.user.id, user.user_role,)
        return prepare_response(message=constants.UNAUTHORIZED_ROLE, status=status.HTTP_403_FORBIDDEN)
    properties_qs = properties_qs.select_related("owner__user", "company", "property").prefetch_related(
        Prefetch("lease_details", queryset=Lease.objects.select_related("tenant__user"))
    ).distinct()
    if tenancy_status == constants.VACANT:
        properties_qs = properties_qs.filter(is_occupied=False)
    elif tenancy_status == constants.OCCUPIED:
        properties_qs = properties_qs.filter(is_occupied=True)
    if search:
        properties_qs = properties_qs.filter(
            Q(unit_name__icontains=search) |
            Q(property__property_name__icontains=search) |
            Q(owner__user__first_name__icontains=search) |
            Q(owner__user__last_name__icontains=search) |
            Q(company__company_name__icontains=search) |
            Q(lease_details__tenant__user__first_name__icontains=search) |
            Q(lease_details__tenant__user__last_name__icontains=search)
        ).distinct()

    now = timezone.now()
    if agreement_expiration == constants.EXPIRED:
        properties_qs = properties_qs.filter(
            lease_details__lease_end_date__lt=now
        )
    elif agreement_expiration == constants.ABOUT_TO_EXPIRE:
        properties_qs = properties_qs.filter(
            lease_details__lease_end_date__gte=now,
            lease_details__lease_end_date__lte=now + timedelta(days=30)
        )
    elif agreement_expiration == constants.ONGOING:
        properties_qs = properties_qs.filter(
            lease_details__lease_end_date__gt=now + timedelta(days=30)
        )

    paginator = Paginator(properties_qs, limit)
    try:
        properties_page = paginator.page(page)
    except EmptyPage:
        properties_page = paginator.page(paginator.num_pages)
    data = []
    for prop in properties_page:
        lease_obj = prop.lease_details.first()
        tenant_name = lease_obj.tenant.user.get_full_name() if lease_obj else None
        owner_name = prop.owner.user.get_full_name() if prop.owner else None
        lease_status = get_lease_status(lease_obj)
        image_data = get_property_images(prop.property.id, single=True) if prop.property else {"images": []}
        property_image = image_data["images"][0]["data"] if image_data.get("images") else None
        tenant_profile_image = lease_obj.tenant.profile_image if lease_obj else None

        data.append({
            "property_unit_id": prop.id,
            "property_name": prop.property.property_name if prop.property else None,
            "unit_name": prop.unit_name,
            "property_code":prop.property.Property_code if prop.property else None,
            "tenant_name": tenant_name,
            "tenant_profile_image": tenant_profile_image,
            "owner_name": owner_name,
            "company_name": prop.company.company_name if prop.company else None,
            "tenancy_status": "occupied" if prop.is_occupied else "vacant",
            "dimension":prop.dimension,
            "agreement_expiration":lease_status,
             "property_image": property_image,

        })

    pagination_meta = {
        "current_page": properties_page.number,
        "limit": limit,
        "total_records": paginator.count,
        "total_pages": paginator.num_pages
    }
    logger.info(
        "PROPERTY_TABLE_FETCHED | user_id=%d | total_records=%d | page=%d",
        request.user.id, paginator.count, page, )

    return prepare_response(
        content=data,
        message=constants.PROPERTIES_FETCHED,
        pagination=pagination_meta,
        status=status.HTTP_200_OK
    )


def parent_property_view(request):
    if request.method == "GET":
        property_id = request.GET.get("id")

        if not property_id:
            logger.warning(
                "PARENT_PROPERTY_FETCH_FAILED | reason=PROPERTY_ID_REQUIRED")
            return prepare_response(
                message=constants.PROPERTY_ID_REQUIRED,
                status=status.HTTP_400_BAD_REQUEST
            )

        prop = Property.objects.filter(id=property_id).first()

        if not prop:
            logger.warning(
                "PARENT_PROPERTY_FETCH_FAILED | property_id=%s | reason=PROPERTY_NOT_FOUND",property_id,)
            return prepare_response(
                message=constants.PROPERTY_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        data = {
            "id": prop.id,
            "property_name": prop.property_name,
            "no_of_blocks": prop.no_of_blocks,
            "no_of_units": prop.no_of_units,
            "property_type": {
                "key": prop.property_type,
                "value": prop.get_property_type_display()
            },
            "land_area": prop.land_area,
            "land_area_unit": prop.land_area_unit,
            "land_dm_no": prop.land_dm_no,
            "plot_no": prop.plot_no,
            "dewa_no": prop.dewa_no,
            "address_line_1": prop.address_line_1,
            "address_line_2": prop.address_line_2,
            "landmark": prop.landmark,
            "pincode": prop.pincode,
            "latitude": prop.latitude,
            "longitude": prop.longitude,
            "map_address": prop.map_address,
            "property_pmc": {
                "id": prop.property_pmc.id if prop.property_pmc else None,
                "name": prop.property_pmc.company_name if prop.property_pmc else None,
            }
        }
        logger.info(
            "PARENT_PROPERTY_FETCHED | property_id=%d | property_name=%s",
            prop.id, prop.property_name, )

        return prepare_response(
            content=data,
            message=constants.PROPERTY_FETCH_SUCCESS,
            status=status.HTTP_200_OK
        )

    else:
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


@is_request_authenticated
def export_property_table_csv(request):
    if request.method != "GET":
        return prepare_response(
            message=constants.INVALID_REQUEST,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    try:
        user = request.user
        search = request.GET.get("search", "").strip()
        if user.user_role == constants.OWNER:
            properties_qs = Unit.objects.filter(owner=user)

        elif user.user_role == constants.COMPANY_USER:
            company = PropertyManagmentCompany.objects.filter(company_user=user).first()
            if not company:
                logger.warning(
                    "PROPERTY_EXPORT_FAILED | user_id=%d | reason=COMPANY_NOT_FOUND",
                    request.user.id,)
                return prepare_response(message=constants.COMPANY_NOT_FOUND, status=status.HTTP_400_BAD_REQUEST)
            properties_qs = Unit.objects.filter(company=company)

        elif user.user_role == constants.TENANT:
            properties_qs = Unit.objects.filter(
                lease_details__tenant=user
            )

        else:
            logger.warning(
                "PROPERTY_EXPORT_FAILED | user_id=%d | reason=UNAUTHORIZED_ROLE",request.user.id,)
            return prepare_response(
                message=constants.UNAUTHORIZED_USER_ROLE,
                status=status.HTTP_403_FORBIDDEN
            )

        properties_qs = properties_qs.select_related(
            "owner__user", "company", "property"
        ).prefetch_related(
            Prefetch(
                "lease_details",
                queryset=Lease.objects.select_related("tenant__user")
            )
        ).distinct()
        if search:
            properties_qs = properties_qs.filter(
                Q(unit_name__icontains=search) |
                Q(property__property_name__icontains=search) |
                Q(owner__user__first_name__icontains=search) |
                Q(owner__user__last_name__icontains=search) |
                Q(company__company_name__icontains=search) |
                Q(lease_details__tenant__user__first_name__icontains=search) |
                Q(lease_details__tenant__user__last_name__icontains=search)
            ).distinct()

        export_data = []


        if user.user_role == constants.OWNER:
            field_names = [
                "Property Name",
                "Tenant Name",
                "Agreement Expiration",
                "Dimension",
                "PropertyManagmentCompany Name"
            ]

            for prop in properties_qs:
                lease = prop.lease_details.first()
                export_data.append({
                    "Property Name": prop.property.property_name if prop.property else "",
                    "Tenant Name": lease.tenant.user.get_full_name() if lease else "",
                    "Agreement Expiration": get_lease_status(lease),
                    "Dimension": prop.area_of_property,
                    "PropertyManagmentCompany Name": prop.company.company_name if prop.company else ""
                })


        elif user.user_role == constants.COMPANY_USER:
            field_names = [
                "Property ID",
                "Property Name",
                "Tenant Name",
                "Owner Name",
                "Tenancy Status"
            ]

            for prop in properties_qs:
                lease = prop.lease_details.first()
                export_data.append({
                    "Property ID": prop.id,
                    "Property Name": prop.property.property_name if prop.property else "",
                    "Tenant Name": lease.tenant.user.get_full_name() if lease else "",
                    "Owner Name": prop.owner.user.get_full_name() if prop.owner else "",
                    "Tenancy Status": "Occupied" if prop.is_occupied else "Available"
                })


        elif user.user_role == constants.TENANT:
            field_names = [
                "Property Name",
                "Owner Name",
                "Tenancy Status",
                "Dimension",
                "PropertyManagmentCompany Name"
            ]

            for prop in properties_qs:
                export_data.append({
                    "Property Name": prop.property.property_name if prop.property else "",
                    "Owner Name": prop.owner.user.get_full_name() if prop.owner else "",
                    "Tenancy Status": "Occupied" if prop.is_occupied else "Available",
                    "Dimension": prop.area_of_property,
                    "PropertyManagmentCompany Name": prop.company.company_name if prop.company else ""
                })
        logger.info(
            "PROPERTY_EXPORTED | user_id=%d | records=%d",request.user.id, len(export_data),)
        return export_to_csv(
            filename="property_table",
            field_names=field_names,
            data_list=export_data
        )

    except Exception as e:
        logger.exception(
            "PROPERTY_EXPORT_EXCEPTION | user_id=%d | error=%s",
            request.user.id, str(e) )
        return prepare_response(
            message=f"Error exporting CSV: {str(e)}",
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@is_request_authenticated
def toggle_property_interest(request):
    if request.method != "PUT":
        logger.warning(
            "PROPERTY_INTEREST_UPDATE_FAILED | user_id=%d | method=%s | reason=INVALID_REQUEST_METHOD",
            request.user.id, request.method,)
        return prepare_response(
            message=constants.INVALID_REQUEST,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    try:
        data = json.loads(request.body)

        property_unit_id = data.get("property_unit_id")
        is_interested = data.get("is_interested")

        if property_unit_id  is None or is_interested is None:
            logger.warning(
                "PROPERTY_INTEREST_UPDATE_FAILED | user_id=%d | reason=INVALID_DATA",
                request.user.id,)
            return prepare_response(
                message=constants.INVALID_DATA,
                status=status.HTTP_400_BAD_REQUEST
            )

        tenant = request.user

        if tenant.user_role != constants.TENANT:
            logger.warning(
                "PROPERTY_INTEREST_UPDATE_DENIED | user_id=%d | reason=ONLY_TENANT_ALLOWED",
                request.user.id,)
            return prepare_response(
                message=constants.ONLY_TENANT_ALLOWED,
                status=status.HTTP_403_FORBIDDEN
            )

        property_unit = Unit.objects.filter(
            id=property_unit_id,
            is_active=True
        ).first()

        if not property_unit:
            logger.warning(
                "PROPERTY_INTEREST_UPDATE_FAILED | user_id=%d | property_unit_id=%s | reason=PROPERTY_UNIT_NOT_FOUND",
                request.user.id, property_unit_id,)
            return prepare_response(
                message=constants.PROPERTY_UNIT_NOT_FOUND,
                status=status.HTTP_404_NOT_FOUND
            )

        with transaction.atomic():
            interest_obj, created = PropertyInterest.objects.update_or_create(
                property_unit=property_unit,
                tenant=tenant,
                defaults={
                    "is_active": is_interested,
                    "created_by": tenant.user
                }
            )
            if is_interested:
                PropertyInterest.objects.filter(
                    id=interest_obj.id
                ).update(created=timezone.now())
        logger.info(
            "PROPERTY_INTEREST_UPDATED | user_id=%d | property_unit_id=%d | interested=%s",
            request.user.id, property_unit.id, is_interested,)
        return prepare_response(
            message=constants.INTEREST_UPDATED_SUCCESS,
            status=status.HTTP_200_OK,
            content={
                "property_unit": property_unit.unit_name,
                "is_interested": is_interested
            }
        )

    except Exception as e:
        logger.exception(
            "PROPERTY_INTEREST_EXCEPTION | user_id=%d",request.user.id,)
        print("Interest API Error:", e)
        return prepare_response(
            message=constants.INTERNAL_SERVER_ERROR,
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def company_list(request):
    if request.method != "GET":
        logger.warning(
            "COMPANY_LIST_INVALID_METHOD | method=%s",request.method,)
        return prepare_response(
            message=constants.INVALID_REQUEST_METHOD,
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    search = request.GET.get("search", "").strip()
    companies = PropertyManagmentCompany.objects.filter(is_active=True).order_by("name")
    if search:
        companies = companies.filter(name__icontains=search)
    content = [{"key": c.id, "value": c.name, "code": c.code} for c in companies]
    logger.info(
        "COMPANY_LIST_FETCHED | total_records=%d",companies.count(),)
    return prepare_response(content=content, status=status.HTTP_200_OK)


# HELPER: Parse sheet — row 1 = description (skip), row 2 = headers
def _parse_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[1]]
    result = []
    for row in rows[2:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


# HELPER: Safe value getter
def _val(data, key, default=None):
    v = data.get(key)
    if v is None:
        return default
    if isinstance(v, str):
        v = v.strip()
        return v if v else default
    return v

# MAIN VIEW: Bulk upload Property + Block + Unit from Excel (base64)
@is_request_authenticated
def bulk_upload_property_excel(request):
    if request.method == "POST":
        user_profile = request.user

        # PMC check
        pm_profile = PropertyManager.objects.filter(pk=user_profile.pk).select_related("company").first()
        pmc = pm_profile.company if pm_profile else None
        if not pmc:
            logger.warning(
                "BULK_UPLOAD_DENIED | user_id=%s | reason=NOT_VERIFIED_PROPERTY_MANAGER",request.user.id)
            return prepare_response(message=constants.NOT_VERIFIED_PROPERTY_MANAGER, status=status.HTTP_403_FORBIDDEN)

        # Base64 file decode
        try:
            body = json.loads(request.body)
        except Exception:
            return prepare_response(message="Invalid JSON body",status=status.HTTP_400_BAD_REQUEST)

        file_base64 = body.get("file")
        if not file_base64:
            return prepare_response(message="Excel file is required (base64 encoded)",status=status.HTTP_400_BAD_REQUEST)

        try:
            if "," in file_base64:
                file_base64 = file_base64.split(",", 1)[1]

            file_bytes = base64.b64decode(file_base64)
            file_stream = io.BytesIO(file_bytes)
            wb = openpyxl.load_workbook(file_stream, data_only=True)
        except Exception as e:
            return prepare_response(
                message=f"Invalid Excel file: {str(e)}",
                status=status.HTTP_400_BAD_REQUEST
            )

        # Sheet check 
        sheet_names = wb.sheetnames
        data_sheets = [s for s in sheet_names if s.lower() != "choices"]
        if len(data_sheets) < 3:
            return prepare_response(
                message=f"Excel must have Property, Block, Unit sheets. Found: {data_sheets}",
                status=status.HTTP_400_BAD_REQUEST
            )

        property_rows = _parse_sheet(wb[data_sheets[0]])
        block_rows    = _parse_sheet(wb[data_sheets[1]])
        unit_rows     = _parse_sheet(wb[data_sheets[2]])
        errors = []
        summary = {
            "properties_created": 0,
            "blocks_created": 0,
            "units_created": 0,
        }

        # ── STEP 1: Create Properties ─────────────────────────────
        property_map = {}
        for i, row in enumerate(property_rows, start=3):
            property_name = _val(row, "property_name_*") or _val(row, "property_name")
            if not property_name:
                errors.append({"sheet": "Property", "row": i, "error": "property_name is required"})
                continue
            if property_name.lower() in property_map:
                errors.append({
                    "sheet": "Property", "row": i,
                    "error": f"Duplicate property_name '{property_name}' in Excel"
                })
                continue

            existing = Property.objects.filter(property_name__iexact=property_name, pmc=pmc).first()
            if existing:
                property_map[property_name.lower()] = existing
                errors.append({
                    "sheet": "Property", "row": i,
                    "error": f"Property '{property_name}' already exists — skipped, using existing"
                })
                continue

            try:
                prop = Property.objects.create(
                    created_by=user_profile.user,
                    property_name=property_name,
                    #property_type=_val(row, "property_type_*") or _val(row, "property_type") or constants.APARTMENT,
                    status=_val(row, "status") or "DRAFT",
                    no_of_blocks=_val(row, "no_of_blocks_*") or _val(row, "no_of_blocks") or 0,
                    no_of_units=_val(row, "no_of_units_*") or _val(row, "no_of_units") or 0,
                    land_area=_val(row, "land_area"),
                    land_area_unit=_val(row, "land_area_unit") or constants.SQ_FT,
                    land_dm_no=_val(row, "land_dm_no"),
                    plot_no=_val(row, "plot_no"),
                    dewa_no=_val(row, "dewa_no"),
                    address_line_1=_val(row, "address_line_1_*") or _val(row, "address_line_1") or "",
                    address_line_2=_val(row, "address_line_2") or "",
                    landmark=_val(row, "landmark") or "",
                    pincode=_val(row, "pincode") or "",
                    latitude=_val(row, "latitude"),
                    longitude=_val(row, "longitude"),
                    map_address=_val(row, "map_address"),
                    approx_rent=_val(row, "approx_rent"),
                    pmc=pmc,
                )
                property_types = (_val(row, "property_type_*")or _val(row, "property_type")or constants.APARTMENT)
                if property_types:
                    codes = [x.strip() for x in re.split(r"[,/]", str(property_types)) if x.strip()]
                    types = PropertyType.objects.filter(code__in=codes)
                    prop.property_type.set(types)
                property_map[property_name.lower()] = prop
                summary["properties_created"] += 1
                logger.info(
                    "BULK_PROPERTY_CREATED | user_id=%s | property_id=%s | name=%s",
                    request.user.id, prop.id, prop.property_name
                )
            except Exception as e:
                errors.append({"sheet": "Property", "row": i, "error": str(e)})

        # ── STEP 2: Create Blocks ─────────────────────────────────
        # if not block_name skip — unit directly link to property 
        block_map = {}

        for i, row in enumerate(block_rows, start=3):
            property_name = _val(row, "property_name_*") or _val(row, "property_name")
            block_name    = _val(row, "block_name_*") or _val(row, "block_name")

            if not property_name:
                errors.append({"sheet": "Block", "row": i, "error": "property_name is required"})
                continue

            if not block_name:
                continue

            prop = property_map.get(property_name.lower())
            if not prop:
                prop = Property.objects.filter(property_name__iexact=property_name, pmc=pmc).first()

            if not prop:
                errors.append({
                    "sheet": "Block", "row": i,
                    "error": f"Property '{property_name}' not found"
                })
                continue
            block_key = f"{property_name.lower()}|{block_name.lower()}"

            if block_key in block_map:
                errors.append({
                    "sheet": "Block", "row": i,
                    "error": f"Duplicate block_name '{block_name}' for property '{property_name}'"
                })
                continue

            existing_block = PropertyBlocks.objects.filter(block_name__iexact=block_name,property=prop).first()
            if existing_block:
                block_map[block_key] = existing_block
                errors.append({
                    "sheet": "Block", "row": i,
                    "error": f"Block '{block_name}' already exists for '{property_name}' — skipped, using existing"
                })
                continue

            try:
                b = PropertyBlocks.objects.create(
                    created_by=user_profile.user,
                    property=prop,
                    block_name=block_name,
                    makani_no=_val(row, "makani_no") or "",
                    no_of_floors=_val(row, "no_of_floors_*") or _val(row, "no_of_floors") or 0,
                    no_of_parking=_val(row, "no_of_parking_*") or _val(row, "no_of_parking") or 0,
                    no_of_units=_val(row, "no_of_units_*") or _val(row, "no_of_units") or 0,
                )
                block_map[block_key] = b
                summary["blocks_created"] += 1
                logger.info(
                    "BULK_BLOCK_CREATED | user_id=%s | block_id=%s | name=%s | property=%s",
                    request.user.id, b.id, b.block_name, property_name
                )
            except Exception as e:
                errors.append({"sheet": "Block", "row": i, "error": str(e)})

        # ── STEP 3: Create Units ──────────────────────────────────
        # block_name  → property_block_tower=block, property=None
        # block_name not → property_block_tower=None, property=prop
        for i, row in enumerate(unit_rows, start=3):
            property_name = _val(row, "property_name_*") or _val(row, "property_name")
            block_name    = _val(row, "block_name_*") or _val(row, "block_name")
            unit_name     = _val(row, "unit_name_*") or _val(row, "unit_name")

            if not property_name:
                errors.append({"sheet": "Unit", "row": i, "error": "property_name is required"})
                continue
            if not unit_name:
                errors.append({"sheet": "Unit", "row": i, "error": "unit_name is required"})
                continue

            if block_name:
                # block_name — link to block 
                block_key = f"{property_name.lower()}|{block_name.lower()}"
                block = block_map.get(block_key)

                if not block:
                    block = PropertyBlocks.objects.filter(
                        block_name__iexact=block_name,
                        property__property_name__iexact=property_name,
                        property__pmc=pmc
                    ).first()

                if not block:
                    errors.append({
                        "sheet": "Unit", "row": i,
                        "error": f"Block '{block_name}' under property '{property_name}' not found"
                    })
                    continue

                # Unit duplicate check — same block has same unit_name
                existing_unit = Unit.objects.filter(
                    unit_name__iexact=str(unit_name),
                    property_block_tower=block
                ).first()
                if existing_unit:
                    errors.append({
                        "sheet": "Unit", "row": i,
                        "error": f"Unit '{unit_name}' already exists in block '{block_name}' — skipped"
                    })
                    continue

                link_kwargs = {"property_block_tower": block, "parent_property": block.property, }

            else:
                # block_name not — directly property link 
                prop = property_map.get(property_name.lower())
                if not prop:
                    prop = Property.objects.filter(property_name__iexact=property_name,pmc=pmc).first()

                if not prop:
                    errors.append({
                        "sheet": "Unit", "row": i,
                        "error": f"Property '{property_name}' not found"
                    })
                    continue

                # Unit duplicate check — same property in same unit_name (no block)
                existing_unit = Unit.objects.filter(unit_name__iexact=str(unit_name),parent_property=prop,property_block_tower=None).first()
                if existing_unit:
                    errors.append({
                        "sheet": "Unit", "row": i,
                        "error": f"Unit '{unit_name}' already exists in property '{property_name}' — skipped"
                    })
                    continue
                link_kwargs = {
                    "property_block_tower": None,
                    "parent_property": prop,
                }

            try:
                u = Unit.objects.create(
                    created_by=user_profile.user,
                    unit_name=str(unit_name),
                    unit_size=_val(row, "unit_size"),
                    area=_val(row, "area"),
                    dm_no=_val(row, "dm_no"),
                    no_of_bedrooms=_val(row, "no_of_bedrooms"),
                    floor_no=_val(row, "floor_no"),
                    parking_no=str(_val(row, "parking_no")) if _val(row, "parking_no") else None,
                    no_of_balcony=_val(row, "no_of_balcony"),
                    land_no=_val(row, "land_no"),
                    unit_usage=_val(row, "unit_usage"),
                    unit_type=_val(row, "unit_type"),
                    sub_type=_val(row, "sub_type"),
                    makani_no=str(_val(row, "makani_no")) if _val(row, "makani_no") else None,
                    dewa_no=_val(row, "dewa_no"),
                    rent=_val(row, "rent"),
                    security_deposit=_val(row, "security_deposit"),
                    booking_amount=_val(row, "booking_amount"),
                    maintenance_charges=_val(row, "maintenance_charges"),
                    cycle=_val(row, "cycle"),
                    notice_period=str(_val(row, "notice_period")) if _val(row, "notice_period") else None,
                    commission_percent=_val(row, "commission_percent"),
                    **link_kwargs,
                )
                summary["units_created"] += 1
                logger.info(
                    "BULK_UNIT_CREATED | user_id=%s | unit_id=%s | name=%s | block=%s | property=%s",
                    request.user.id, u.id, u.unit_name,
                    block_name or "None",
                    property_name
                )
            except Exception as e:
                errors.append({"sheet": "Unit", "row": i, "error": str(e)})

        # ── Response ──────────────────────────────────────────────
        audit_logs(request, f"Bulk upload: {summary}", constants.CREATED)

        if errors and summary["properties_created"] == 0 and summary["blocks_created"] == 0 and summary["units_created"] == 0:
            return prepare_response(
                message="Bulk upload failed",
                content={"errors": errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        return prepare_response(
            message="Bulk upload completed",
            content={
                "summary": summary,
                "errors": errors
            },
            status=status.HTTP_201_CREATED
        )
